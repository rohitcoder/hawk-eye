import jmespath
from rich.console import Console 
from rich.table import Table
import json, requests, argparse, yaml, re, datetime, os, subprocess, platform, hashlib
from tinydb import TinyDB, Query
import numpy as np
import pytesseract
from PIL import Image, ImageEnhance
from docx import Document
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
import PyPDF2
import patoolib
import tempfile
import shutil
import os, cv2
import tarfile
import pkg_resources
from concurrent.futures import ProcessPoolExecutor


data_sources = ['s3', 'mysql', 'redis', 'firebase', 'gcs', 'fs', 'postgresql', 'mongodb', 'slack', 'couchdb', 'gdrive', 'gdrive_workspace', 'text']
data_sources_option = ['all'] + data_sources

def parse_args(args=None):
    version = pkg_resources.require("hawk_scanner")[0].version 
    parser = argparse.ArgumentParser(description='🦅 A powerful scanner to scan your Filesystem, S3, MySQL, PostgreSQL, MongoDB, Redis, Google Cloud Storage and Firebase storage for PII and sensitive data.')
    parser.add_argument('command', nargs='?', choices=data_sources_option, help='Command to execute')
    parser.add_argument('--connection', action='store', help='YAML Connection file path')
    parser.add_argument('--connection-json', type=str, help='Connection details in JSON format, useful for passing connection info directly as CLI Input')
    parser.add_argument('--fingerprint', action='store', help='Override YAML fingerprint file path')
    parser.add_argument('--json', help='Save output to a json file')
    parser.add_argument('--stdout', action='store_true', help='Print output to stdout in JSON format')
    parser.add_argument('--quiet', action='store_true', help='Print only the results')
    parser.add_argument('--debug', action='store_true', help='Enable debug mode')
    parser.add_argument('--no-write', action='store_true', help='Do not write previous alerts to file, this may flood you with duplicate alerts')
    parser.add_argument('--shutup', action='store_true', help='Suppress the Hawk Eye banner 🫣', default=False)
    parser.add_argument('--version', action='version', version='%(prog)s v' + version) 
    parser.add_argument('--hawk-thuu', action='store_true', help="Delete all spitted files during testing phase forcefully")
    return parser.parse_args(args)
    
console = Console()

def calculate_msg_hash(msg):
    return hashlib.sha256(msg.encode()).hexdigest()

def print_info(args, message):
    if not args.quiet:
        console.print(f"[yellow][INFO][/yellow] {str(message)}")

def print_debug(args, message):
    if args and type(args) == argparse.Namespace and args.debug and not args.quiet:
        try:
            console.print(f"[blue][DEBUG][/blue] {str(message)}")
        except Exception as e:
            pass

def print_error(args, message):
    if not args.quiet:
        console.print(f"[bold red]❌ {message}")

def print_success(args, message):
    if not args.quiet:
        console.print(f"[bold green]✅ {message}")

def get_file_owner(file_path):
    owner_name = ""

    # Determine the current operating system
    system = platform.system()

    if system == "Windows":
        try:
            # Run the 'dir /q' command and capture its output
            result = subprocess.check_output(['dir', '/q', file_path], shell=True, text=True)
            # Extract the line containing the file information
            lines = result.splitlines()
            file_name = os.path.basename(file_path)
            if len(lines) >= 6:
                for line in lines:
                    if file_name in line:
                        exploded_line = line.split(' ')
                        owner_name = exploded_line[-2]
        except subprocess.CalledProcessError as e:
            owner_name = ""
    else:
        try:
            from pwd import getpwuid
            # Use the 'os.stat()' method to get the file owner on non-Windows systems
            file_stat = os.stat(file_path)
            owner_name = file_stat.st_uid  # You can also use pwd.getpwuid(file_stat.st_uid).pw_name to get the username
            owner_name = getpwuid(owner_name).pw_name + " (" + str(owner_name) + ")"
        except OSError as e:
            owner_name = ""

    return owner_name

def RedactData(input_string):
    if len(input_string) < 3:
        return input_string

    # Calculate the number of characters to redact in the middle (half of the length)
    redact_count = len(input_string) // 2

    # Split the input string into two parts: before and after the middle
    middle_start = len(input_string) // 2 - redact_count // 2
    middle_end = len(input_string) // 2 + redact_count // 2

    before_middle = input_string[:middle_start]
    middle = input_string[middle_start:middle_end]
    after_middle = input_string[middle_end:]

    # Redact the middle part
    redacted_middle = "*" * len(middle)

    # Concatenate the parts back together
    redacted_string = before_middle + redacted_middle + after_middle

    return redacted_string

def get_connection(args):
    if args.connection:
        if os.path.exists(args.connection):
            with open(args.connection, 'r') as file:
                connections = yaml.safe_load(file)
                return connections
        else:
            print_error(args, f"Connection file not found: {args.connection}")
            exit(1)
    elif args.connection_json:
        try:
            connections = json.loads(args.connection_json)
            return connections
        except json.JSONDecodeError as e:
            print_error(args, f"Error parsing JSON: {e}")
            exit(1)
    else:
        print_error(args, "Please provide a connection file using --connection flag or connection details using --connection-json flag")
        exit(1)

def get_fingerprint_file(args=None):
    if args and type(args) == argparse.Namespace and args.fingerprint:
        if os.path.exists(args.fingerprint):
            with open(args.fingerprint, 'r') as file:
                return yaml.safe_load(file)
        else:
            if args:
                print_error(args, f"Fingerprint file not found: {args.fingerprint}")
            exit(1)
    elif args and type(args) == dict and 'fingerprint' in args:
        return args['fingerprint']
    else:
        file_path = "https://github.com/rohitcoder/hawk-eye/raw/main/fingerprint.yml"
        try:
            response = requests.get(file_path, timeout=10)
            if args:
                print_info(args, f"Downloading default fingerprint.yml from {file_path}")
            if response.status_code == 200:
                with open('fingerprint.yml', 'wb') as file:
                    file.write(response.content)
                return yaml.safe_load(response.content)
            else:
                if args:
                    print_error(args, f"Unable to download default fingerprint.yml please provide your own fingerprint file using --fingerprint flag")
                exit(1)
        except Exception as e:
            if args:
                print_error(args, f"Unable to download default fingerprint.yml please provide your own fingerprint file using --fingerprint flag")
            exit(1)

def print_banner(args):
    line1 = "+ ================================================== +"
    line2 = "+ [bold yellow]H[/bold yellow].[bold yellow]A[/bold yellow].[bold yellow]W[/bold yellow].[bold yellow]K[/bold yellow] [bold yellow]Eye[/bold yellow] - [bold blue]Highly Advanced Watchful Keeper Eye[/bold blue] +"
    line3 = "+ ================================================== +"
    line4 = ""
    line5 = "Hunt for Secrets & PII Data, like never before!"
    line6 = "A Tool by [bold red]Rohit Kumar (@rohitcoder)[/bold red]"

    banner = f"""
                              ⢀⣀⣀⣀⣀⣤⠶⣒⣛⣛⡲⠶⣶⣶⠤⣤⣀⡀              
                            ⣠⠴⢻⣭⣶⣶⣶⣶⣶⣦⣽⣿⣿⣿⣷⣾⣥⣝⣶⣭⣗⣲⡕⣠⡀         
                          ⢀⣨⣷⣺⣵⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⣵⣝⣦⡀       
                        ⣠⡾⡫⢵⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡟⣿⣿⡿⣿⣿⣞⣿⡄      
                       ⢼⣟⣯⣾⣿⣿⣿⣿⣿⡍⠉⠙⠛⠟⠿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡌⣿⣷⢸⣿⣿⣾⣿      
                      ⢠⢫⣾⣿⣿⣿⣿⣿⠋⠁ ⠠     ⠉⠻⢿⣿⣿⣿⣿⣿⠟⢀⣿⠟⣘⣿⣿⣿⣿      
                     ⣰⣏⣿⣿⣿⣿⣿⣿⣿⣇ ⠠⡖⠒⣲⣶⣶⣶⣦⣤⣄⡛⠛⠛⢛⣁⣤⡽⠿⠋⠉⠉ ⠙⠛⠒⠦⣄   
                     ⢿⡅⣼⣿⣿⣿⣿⣿⣿⣿⠗ ⠱⡀⢿⣿⣤⣧⡼⢻⢻⣿⠟⠉⢻⠟⠁⡰⠟⠁       ⠈⢳⡄ 
                    ⢀⡾⣼⣿⣿⣿⣿⣿⣿⣿⣿⣿  ⠈⠓⠭⠿⠭⠄⡊⠤⠊⠎⢢⠏              ⢳ 
                    ⣼⢱⣿⢿⣿⣿⣿⣿⣿⣿⣿⣿⣄⡀    ⡀⢀    ⡜⢠⢠⢀⢀          ⠈⣾⡄	 {line1}
                    ⣿⣨⢅⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣷⡶⠄⣀⣀⡀  ⡠⢊⣀   ⠈           ⢿⡇	 {line2}
                   ⢀⡴⣽⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⠟⠁ ⡘⠠⠋⢀⠤⢒⣒⣒⣒⠢⠤⢤⣄⣀⣀⣀⡀  ⣼⠁   {line3}
                   ⣎⠾⣻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠆   ⠱⢄⢘⣡⠔⠶⢴⣒⣢⣤⡤⠖⠛⡏⠉⠉⠉⢣⢀⡟    {line4}
                  ⠸⣴⣳⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⠿⠛⠁ ⠐  ⢀⠒⣭⠅    ⠰⠌⠁ ⡀⣷    ⠛     {line5}
                 ⢀⡼⣱⣿⣿⣿⣿⣿⣿⣿⡿⣻⣿⣿⣷⡦⠄    ⠈ ⠈⠁ ⠂    ⠈  ⢀⣾⣧⣿⠁         {line6} 
                ⣠⣿⢃⢋⣽⣿⣿⣿⣿⣿⣿⠵⣿⡿⣿⣿⣶⡿⣃⣀⣤⣴⣇⣠⣾⡇       ⢠⡗⣸⣿⣿⣿⠆      
              ⢀⣾⣿⢟⣵⣿⣿⣿⣿⣿⣿⣿⣿⣿⣟⣶⣿⣿⣿⣿⣿⢿⣿⣿⢿⣿⡟  ⢀⡄   ⢀⣾⣿⣿⣿⣿⣯       
             ⢀⣿⣿⣯⣶⢖⣸⣟⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣃⣿⣭⣾⣿⣯⣤⣴⣾⣿⢇⢌⣾⡆⣼⣿⣿⣿⣿⣏⢃⠃      
            ⠠⣫⠾⣻⣿⢿⣿⠫⠾⠿⠿⣿⡛⣫⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡟⣷⣄⡀     
            ⠐⠁⣼⠟⢡⣿⠿⣿⣿⣿⡿⢟⣡⢖⣾⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣛⣻⢹⣇⡀     
             ⠐⠁⢀⠟⢁⣾⣿⠿⣿⣾⣿⢯⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⡿⡿⣿⢓⣛⣿⣿⡿⢆⡀   
               ⠊⢀⡞⠉ ⣴⠟⣻⣷⣶⣶⣾⣭⣽⣿⣭⣵⡾⣻⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣿⣉⠿⣩⣣⣿⣷⡞⢟⠟⣿⢛⠆⠈⠂  
                ⠈  ⠐⠁⣼⠟⢻⣿⣿⡿⠿⢿⣿⣿⣟⣟⣻⣿⣻⣿⣭⢇⡿⣛⢻⣿⣋⢿⢣⣿⣿⢿⣿⣿⡇   ⠙⠈⠌    
                    ⠈⠁ ⠛⠋ ⣀⣴⣿⣿⣿⡿⢿⣿⣿⣿⣿⣿⣬⣾⣿⣟⣷⣿⣶⣿⠟⠁ ⣿⡟           
                        ⠐⠞⠛⠉    ⣼⠿⠋ ⠉⢹⣿⡿⣿⣿⢿⣿⠏⠃   ⠙            
                                    ⢀⣾⠏⢠⡿⠁⡾⠋                  
                                    ⠈⠁ ⠈                      
"""
    if args.quiet:
        args.shutup = True
    if not args.shutup:
        console.print(banner)

def match_strings(args, content, source='text'):
    redacted = False
    if args and 'connection' in args:
        connections = get_connection(args)
        if 'notify' in connections:
            redacted: bool = connections.get('notify', {}).get('redacted', False)
    
    patterns = get_fingerprint_file(args)
    matched_strings = []

    for pattern_name, pattern_regex in patterns.items():
        if args:
            print_debug(args, f"Matching pattern: {pattern_name}")
        found = {} 
        ## parse pattern_regex as Regex
        complied_regex = re.compile(pattern_regex, re.IGNORECASE)
        if args:
            print_debug(args, f"Regex: {complied_regex}")
            print_debug(args, f"Content: {content}")
        matches = re.findall(complied_regex, content)
        print_debug(args, f"Matches: {matches}")
        found['data_source'] = source
        if matches:
            print_debug(args, f"Found {len(matches)} matches for pattern: {pattern_name}")
            found['pattern_name'] = pattern_name
            redacted_matches = []
            if redacted:
                if args:
                    print_debug(args, f"Redacting matches for pattern: {pattern_name}")
                for match in matches:
                    print_debug(args, f"Redacting match: {match}")
                    redacted_matches.append(RedactData(match))
                found['matches'] = redacted_matches
            else:
                found['matches'] = matches

            if redacted:
                found['sample_text'] = RedactData(content[:50])
            else:
                found['sample_text'] = content[:50]
            if found['matches'] and len(found['matches']) > 0:
                found['matches'] = [x.strip() for x in found['matches']]
                found['matches'] = list(set(found['matches']))
            matched_strings.append(found)
    if args:
        print_debug(args, f"Matched strings: {matched_strings}")
    ## remove duplicates from matches and return
    return matched_strings

def should_exclude_file(args, file_name, exclude_patterns):
    _, extension = os.path.splitext(file_name)
    if extension in exclude_patterns:
        print_debug(args, f"Excluding file: {file_name} because of extension: {extension}")
        return True
    
    for pattern in exclude_patterns:
        if pattern in file_name:
            print_debug(args, f"Excluding file: {file_name} because of pattern: {pattern}")
            return True
    return False

def should_exclude_folder(folder_name, exclude_patterns):
    for pattern in exclude_patterns:
        if pattern in folder_name:
            return True
    return False

def list_all_files_iteratively(args, path, exclude_patterns):
    for root, dirs, files in os.walk(path, topdown=True):
        dirs[:] = [d for d in dirs if not should_exclude_folder(os.path.join(root, d), exclude_patterns)]

        for file in files:
            if not should_exclude_file(args, file, exclude_patterns):
                yield os.path.join(root, file)

def scan_file(file_path, args=None, source=None):
    content = ''
    is_archive = False
    # Check if the file is an image
    if file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
        content = enhance_and_ocr(file_path)
    # Check if the file is a PDF document
    elif file_path.lower().endswith('.pdf'):
        content = read_pdf(args, file_path)
    # Check if the file is an office document (Word, Excel, PowerPoint)
    elif file_path.lower().endswith(('.docx', '.xlsx', '.pptx')):
        content = read_office_document(args, file_path)
    # Check if the file is an archive (zip, rar, tar, tar.gz)
    elif file_path.lower().endswith(('.mp4', '.avi', '.mov', '.mkv')):
        content = read_video(args, file_path)
    elif file_path.lower().endswith(('.zip', '.rar', '.tar', '.tar.gz')):
        ## this is archive, so we need to extract it and find pii from it, and return matched_strings
        matched_strings = find_pii_in_archive(args, file_path, source)
        is_archive = True
    else:
        # For other file types, read content normally
        with open(file_path, 'rb') as file:
            # Attempt to decode using UTF-8, fallback to 'latin-1' if needed
            content = file.read().decode('utf-8', errors='replace')

    if not is_archive:
        matched_strings = match_strings(args, content, source)
    return matched_strings

def read_match_strings(args, file_path, source):
    print_info(args, f"Scanning file: {file_path} for Source: {source}")
    try:
        matched_strings = scan_file(file_path, args, source)
    except Exception as e:
        print_debug(args, f"Error in read_match_strings: {e}")
        matched_strings = []
    return matched_strings

def read_pdf(args, file_path):
    content = ''
    try:
        # Read content from PDF document
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page_num in range(len(pdf_reader.pages)):  # Use len() instead of deprecated numPages
                page = pdf_reader.pages[page_num]
                try:
                    content += page.extract_text()
                except UnicodeDecodeError:
                    # Handle decoding errors by trying a different encoding
                    content += page.extract_text(encoding='latin-1')
    except Exception as e:
        print_debug(args, f"Error in read_pdf: {e}")
    return content

def process_frame(frame, frame_num, args):
    """
    Process a single frame: convert it to grayscale and run OCR.

    :param frame: The video frame to process.
    :param frame_num: The frame number (for debugging).
    :param args: Arguments for debugging or additional settings.
    :return: Extracted text from the frame.
    """
    try:
        gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        # Resize for faster OCR if resolution is not critical
        small_frame = cv2.resize(gray_frame, (gray_frame.shape[1] // 2, gray_frame.shape[0] // 2))

        # Apply thresholding (optional, improves OCR on some images)
        _, thresh_frame = cv2.threshold(small_frame, 150, 255, cv2.THRESH_BINARY)

        # Perform OCR with optimized configuration
        custom_config = r'--oem 3'  # PSM 6 for uniform block of text
        text = pytesseract.image_to_string(thresh_frame, config=custom_config)

        if args.debug:
            print(f"Processed frame {frame_num}")
        
        return text.strip()
    except Exception as e:
        if args.debug:
            print(f"Error processing frame {frame_num}: {e}")
        return ""

def process_frames_parallel(frames, args):
    with ProcessPoolExecutor(max_workers=4) as executor:  # Use multiple processes
        futures = [executor.submit(process_frame, frame, i, args) for i, frame in enumerate(frames)]
        return [future.result() for future in futures]

def read_video(args, file_path, frame_interval=30, max_workers=10):
    """
    Extract text from a video file by applying OCR on its frames.

    :param args: Arguments for debugging or additional settings.
    :param file_path: Path to the video file.
    :param frame_interval: Interval to capture frames (default is every 30 frames).
    :param max_workers: Number of parallel workers (default is 4).
    :return: Extracted text content from video frames.
    """
    content = ''
    try:
        # Open the video file
        cap = cv2.VideoCapture(file_path)

        if not cap.isOpened():
            raise ValueError(f"Cannot open video file: {file_path}")

        frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        frame_rate = cap.get(cv2.CAP_PROP_FPS)
        print_debug(args, f"Processing {frame_count} frames at {frame_rate} FPS")

        futures = []
        processed_frames = 0

        # Create a ThreadPoolExecutor for parallel processing
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            for frame_num in range(frame_count):
                ret, frame = cap.read()
                if not ret:
                    break

                # Process only every `frame_interval`-th frame
                if frame_num % frame_interval == 0:
                    print_debug(args, f"Submitting frame {frame_num}/{frame_count} for processing")
                    # Submit the frame to the thread pool for processing
                    futures.append(executor.submit(process_frame, frame, frame_num, args))
                    processed_frames += 1

            # Wait for all submitted frames to complete and gather the results
            for future in as_completed(futures):
                text = future.result()
                content += text + '\n'

        cap.release()
        print_debug(args, f"Processed {processed_frames} frames out of {frame_count}")

    except Exception as e:
        print_debug(args, f"Error in read_video: {e}")

    return content


def read_office_document(args, file_path):
    content = ''
    try:
        # Check the file type and read content accordingly
        if file_path.lower().endswith('.docx'):
            # Read content from Word document
            doc = Document(file_path)
            for paragraph in doc.paragraphs:
                content += paragraph.text + '\n'
        elif file_path.lower().endswith('.xlsx'):
            # Read content from Excel spreadsheet
            workbook = load_workbook(file_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        content += str(cell.value) + '\n'
        elif file_path.lower().endswith('.pptx'):
            # Read content from PowerPoint presentation
            # You can add specific logic for PowerPoint if needed
            pass
    except Exception as e:
        print_debug(args, f"Error in read_office_document: {e}")
    return content

def find_pii_in_archive(args, file_path, source):
    content = []
    # Create a temporary directory to extract the contents of the archive
    with tempfile.TemporaryDirectory() as tmp_dir:
        # Extract the contents of the archive based on the file extension
        if file_path.lower().endswith('.zip'):
            patoolib.extract_archive(file_path, outdir=tmp_dir)
        elif file_path.lower().endswith('.rar'):
            patoolib.extract_archive(file_path, outdir=tmp_dir)
        elif file_path.lower().endswith('.tar'):
            with tarfile.open(file_path, 'r') as tar:
                tar.extractall(tmp_dir)
        elif file_path.lower().endswith('.tar.gz'):
            with tarfile.open(file_path, 'r:gz') as tar:
                tar.extractall(tmp_dir)
        # Iterate over all files in the temporary directory
        for root, dirs, files in os.walk(tmp_dir):
            for file in files:
                file_path = os.path.join(root, file)
                data = read_match_strings(args, file_path, source)
                for d in data:
                    content.append(d)
        # Clean up the temporary directory
        shutil.rmtree(tmp_dir)
    return content


def getFileData(file_path):
    try:
        # Get file metadata
        file_stat = os.stat(file_path)

        # Get the username of the file's creator (Windows)
        creator_name = get_file_owner(file_path)
        # Convert timestamps to human-readable format
        created_time = datetime.datetime.fromtimestamp(file_stat.st_ctime)
        modified_time = datetime.datetime.fromtimestamp(file_stat.st_mtime)

        # Create a dictionary with the file information
        ## should be dict
        file_info = {
            "creator": creator_name,
            "created_time": created_time.strftime("%Y-%m-%d %H:%M:%S"),
            "modified_time": modified_time.strftime("%Y-%m-%d %H:%M:%S"),
        }
        return file_info

    except FileNotFoundError:
        return json.dumps({"error": "File not found"})
    except Exception as e:
        return json.dumps({"error": str(e)})


def SlackNotify(msg, args):
    connections = get_connection(args)
    if not args.no_write:
        db = TinyDB('previous_alerts.json')
   
    if 'notify' in connections:
        notify_config = connections['notify']
        # Check if suppress_duplicates is set to True
        suppress_duplicates = notify_config.get('suppress_duplicates', False)
        original_msg = msg
        if suppress_duplicates and not args.no_write:
            # Calculate the hash of the message
            ## check if "msg" has "Message Link" in any line, then remove that complete line
            if "Message Link" in msg:
                msg = msg.split("\n")
                msg = [line for line in msg if "Message Link" not in line]
                msg = "\n".join(msg)
            
            msg_hash = calculate_msg_hash(msg)
            # Check if the message hash already exists in the previous alerts database
            alert_query = Query()
            if db.search(alert_query['msg_hash'] == msg_hash):
                print_info(args, "Duplicate message detected. Skipping webhook trigger.")
                return
        
        slack_config = notify_config.get('slack', {})
        webhook_url = slack_config.get('webhook_url', '')
        if webhook_url and webhook_url.startswith('https://hooks.slack.com/services/'):
            try:
                payload = {
                    'text': original_msg,
                }
                headers = {'Content-Type': 'application/json'}
                requests.post(webhook_url, data=json.dumps(payload), headers=headers)
                if suppress_duplicates and not args.no_write:
                    # Store the message hash in the previous alerts database
                    db.insert({'msg_hash': msg_hash})
            except Exception as e:
                print_error(args, f"An error occurred: {str(e)}")

def evaluate_severity(json_data, rules):
    if 'severity_rules' not in rules:
        rules = {
            'severity_rules': {
                'Highest': [
                    {'query': "length(matches) > `20`", 'description': "Detected more than 20 PII or Secrets"},
                ],
                'High': [
                    {'query': "length(matches) > `10` && length(matches) <= `20`", 'description': "Detected more than 10 PII or Secrets"},
                ],
                'Medium': [
                    {'query': "length(matches) > `5` && length(matches) <= `10`", 'description': "Detected more than 5 PII or Secrets"},
                ],
                'Low': [
                    {'query': "length(matches) <= `5`", 'description': "Detected less than 5 PII or Secrets"},
                ],
            }
        }
    
    for severity, conditions in rules['severity_rules'].items():
        for condition in conditions:
            query = condition['query']
            description = condition['description']
            if jmespath.search(query, json_data):
                # Add severity details to the JSON data
                json_data['severity'] = severity
                json_data['severity_description'] = description
                return json_data

    # If no match, add default severity
    json_data['severity'] = "unknown"
    json_data['severity_description'] = "No matching rule found."
    return json_data

def enhance_and_ocr(image_path):
    # Load the image
    original_image = Image.open(image_path)

    # Enhance the image (you can adjust enhancement factors as needed)
    enhanced_image = enhance_image(original_image)

    # Save the enhanced image for reference
    enhanced_image.save("enhanced_image.png")

    # Perform OCR on the enhanced image
    ocr_text = perform_ocr(enhanced_image)
    ## delete the enhanced image
    os.remove("enhanced_image.png")

    return ocr_text

def enhance_image(image):
    # Convert to grayscale
    grayscale_image = image.convert('L')

    # Increase contrast
    contrast_enhancer = ImageEnhance.Contrast(grayscale_image)
    contrast_factor = 2.0  # Adjust as needed
    contrast_enhanced_image = contrast_enhancer.enhance(contrast_factor)

    # Apply thresholding
    threshold_value = 100  # Adjust as needed
    thresholded_image = contrast_enhanced_image.point(lambda x: 0 if x < threshold_value else 255)

    # Reduce noise (optional)
    denoised_image = cv2.fastNlMeansDenoising(np.array(thresholded_image), None, h=10, templateWindowSize=7, searchWindowSize=21)

    return Image.fromarray(denoised_image)

def perform_ocr(image):
    # Use Tesseract OCR
    ocr_text = pytesseract.image_to_string(image)

    return ocr_text

def get_jira_accId(args, email):
    config = get_connection(args)
    jira_config = config.get('notify', {}).get('jira', {})
    server_url = jira_config.get('server_url')
    username = jira_config.get('username')
    api_token = jira_config.get('api_token')
    db = TinyDB('user_ids.json')
    user_query = Query()

    # Check if the accountId is already cached
    cached_user = db.search(user_query.email == email)
    if cached_user:
        print_debug(args, f"Using cached accountId for {email}: {cached_user[0]['accountId']}")
        return cached_user[0]['accountId']

    # Fetch accountId from Jira
    url = f"{server_url}/rest/api/2/user/search?query={email}"
    auth = (username, api_token)
    headers = {"Content-Type": "application/json"}

    response = requests.get(url, auth=auth, headers=headers)
    if response.status_code == 200:
        data = response.json()
        if len(data):
            account_id = data[0]['accountId']
            print_debug(args, f"Found accountId for {email}: {account_id}")

            # Cache the accountId
            db.insert({"email": email, "accountId": account_id})
            return account_id
        else:
            print_debug(args, f"No accountId found for {email}.")
            return None
    else:
        print_debug(args, f"Failed to fetch accountId for {email}: {response.status_code}, {response.text}")
        return None
    
def create_jira_ticket(args, issue_data, message):
    orig_msg = message
    config = get_connection(args)
    if not args.no_write:
        db = TinyDB('previous_alerts.json')
        if 'notify' in config:
            notify_config = config['notify']
            # Check if suppress_duplicates is set to True
            suppress_duplicates = notify_config.get('suppress_duplicates', False)
            if suppress_duplicates and not args.no_write:
                # Calculate the hash of the message
                ## check if "msg" has "Message Link" in any line, then remove that complete line
                if "Message Link" in message:
                    message = message.split("\n")
                    message = [line for line in message if "Message Link" not in line]
                    message = "\n".join(message)
                
                msg_hash = calculate_msg_hash(message)
                # Check if the message hash already exists in the previous alerts database
                alert_query = Query()
                if db.search(alert_query['msg_hash'] == msg_hash):
                    print_info(args, "Duplicate message detected. Skipping ticket creation")
                    return

    """Creates a Jira ticket using the provided configuration and issue data."""
    jira_config = config.get('notify', {}).get('jira', {})

    # Check if Jira is enabled
    if not jira_config.get('username') or jira_config.get('username') == '':
        print_debug(args, "Jira ticket creation is disabled in the configuration.")
        return

    # Extract Jira config details
    server_url = jira_config.get('server_url')
    evaluated_result = evaluate_severity(issue_data, config)
    severity = evaluated_result.get('severity')
    severity_description = evaluated_result.get('severity_description')
    username = jira_config.get('username')
    api_token = jira_config.get('api_token')
    project = jira_config.get('project')
    default_issue_type = jira_config.get('issue_type')
    issue_fields = jira_config.get('issue_fields', {})
    total_matches = len(issue_data.get('matches', []))
    summary = "Found " + str(total_matches) + " " + issue_data.get('pattern_name') + " in " + issue_data.get('data_source')
    description_template = issue_fields.get('description_template', '')
    orig_msg = orig_msg + "\n\n" + "Severity: " + severity + "\n" + "Severity Description: " + severity_description
    description = description_template.format(details=orig_msg, **issue_data)
    print("severity - ", severity)
    payload = {
        "fields": {
            "project": {"key": project},
            "summary": summary,
            "description": description,
            "issuetype": {"name": default_issue_type},
            "priority": {"name": severity},
        }
    }
    
    # Check if the assignee is specified in the configuration
    assignee = jira_config.get('assignee')
    if assignee:
        payload['fields']['assignee'] = {"accountId": get_jira_accId(args, assignee)}
    
    labels = jira_config.get('labels')
    if labels:
        payload['fields']['labels'] = labels

    # Send request to Jira API
    url = f"{server_url}/rest/api/latest/issue"
    auth = (username, api_token)
    headers = {"Content-Type": "application/json"}

    response = requests.post(url, json=payload, auth=auth, headers=headers)
    if response.status_code == 201:
        print_debug(args, f"Jira ticket created successfully: {response.json().get('key')}")
    else:
        print_debug(args, f"Failed to create Jira ticket: {response.status_code} - {response.text}")
